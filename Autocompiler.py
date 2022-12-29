from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, DEFAULT_FONT
import csv
import pyexcel as p
import tkinter as tk
from tkinter import simpledialog, messagebox

#Byung Woong Ko
#08/1/2022
#Strategic Plan Autocompiler

mode = 0


#load excel
def loadWorkBook(fName):
    TWB=load_workbook(filename=fName)
    return TWB

#convert csv to excel
def C2E(csv_file):
    csv_data = []
    with open(csv_file) as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)
    workbook = Workbook()
    sheet = workbook.active
    for row in csv_data:
        sheet.append(row)
    return sheet

def xls2xlsx(a1,a2):
    p.save_book_as(file_name=a1,
               dest_file_name=a2)

#create new excel sheet
def newWorkBook():
    NWB=Workbook()
    return NWB

#create new sheet
def makeSheet(NWB):
    sheet = NWB.active
    return sheet

#make bold
def fBold(id,sheet):
    sheet[id].font= Font(bold=True)

#make keys for excel sheet
def fKeys(array):
    keys=[]
    for k in array:
        if k.value!=None:
            keys.append(k.value)
    return keys

#unique counter
def cntU(str,array,i,j):
    if j==0:
        return 0
    if array[i]==str:
        return 1+cntU(str,array,i+1,j-1)
    else:
        return cntU(str,array,i+1,j-1)

#unique finder
def findU(array):
    set1=set(array)
    if '--' in set1:
        set1.remove('--')
    return len(set1)

#finding sum of uber fares incl tax, tip, and tech fee
def findSumU(a1,a2):
    i=7
    total=0
    while i<len(a1):
        if a1[i]=='Fare':
            total=total+(float(a2[i])*1.1)
        if a1[i]=='Service & Technology Fee':
            total=total+float(a2[i])
        if a1[i]=='Tip':
            total=total+float(a2[i])
        i=i+1
    return total

#finding sum of total stduents enrolled
def findSumE(a1,a2):
    checker=[]
    i=0
    total=0
    while i<len(a1):
        if not a1[i] in checker:
            checker.append(a1[i])
            total=total+float(a2[i])
        i=i+1
    return total

#make Italicent
def fItalic(id,sheet):
    sheet[id].font= Font(italic=True)

#create basic descriptions in excel sheet
def foundation(sheet,year):
    DEFAULT_FONT.size=12
    sheet['A12'].alignment=Alignment(wrap_text=True)
    sheet['A15'].alignment=Alignment(wrap_text=True)
    sheet.column_dimensions['A'].width=74
    sheet.column_dimensions['B'].width=52
    sheet.column_dimensions['C'].width=18
    sheet.column_dimensions['D'].width=18
    sheet.column_dimensions['E'].width=18
    sheet.column_dimensions['F'].width=4
    sheet.column_dimensions['G'].width=18
    sheet.column_dimensions['A'].font=Font(italic=True)
    if year=="":
        year=0
    year=int(year) 
    sheet["C1"] = "Summer "+str(year)
    sheet["D1"] = "Fall " +str(year)
    sheet["E1"] = "Spring " + str(year+1)
    sheet["G1"] ="Total Year"

    sheet["A2"] = "Objective: Develop multiuse, multidisciplinary, and multimodal campus facilities"
    sheet["A12"] = "Objective: Design a campus framework plan to support development and redevelopment opportunities"
    sheet["A15"] = "Objective: Design and implement a transportation system that increase accessibility to greater Boston and facilities travel between Amherst and Newton"
    sheet["A21"] = "Objective: Increase capacity for Mount Ida, campus conferences, and events"
    sheet["A26"] = "Objective: Grow the overall population of students utilizing the Mount Ida Campus"

    
    sheet["B3"] = "Multimodal course enrollments"
    sheet["B4"] = "ugrad"
    sheet["B5"] = "grad"
    
    sheet["B7"] = "Classroom space usd (%)"
    sheet["B8"] = "Rooms listed by registrar"
    sheet["B9"] = "Rooms used"
    sheet["B10"] = "No. departments using classrooms"
    
    sheet["B13"] = "New project initiation"
    
    sheet["B17"] = "Transporation Enrollment (Stud, Fac, Staff)"
    sheet["B18"] = "Transportation cost-efficiency (per person/semster)"
    sheet["B19"] = "Cost"
    
    sheet["B22"] = "Event, Rental and Lease Revenue"
    sheet["B23"] = " - Events"
    sheet["B24"] = " - Space Rental/Use"
    sheet["B25"] = " - Leases"

    sheet["B27"] = "Campus based course enrollment"
    sheet["B28"] = "in person"
    sheet["B29"] = "remote"
    
    sheet["B31"] = "Residential student population (unique residents)"

    sheet["B33"] = "Course sections offered"
    sheet["B34"] = "in person"
    sheet["B35"] = "remote"

 
    fBold("B3",sheet)
    fBold("B7",sheet)
    fBold("B10",sheet)
    fBold("B18",sheet)
    fBold("B22",sheet)
    fBold("B27",sheet)
    fBold("B31",sheet)
    fBold("B33",sheet)
    fBold("C3",sheet)
    fBold("C7",sheet)
    fBold("C10",sheet)
    fBold("C18",sheet)
    fBold("C22",sheet)
    fBold("C27",sheet)
    fBold("C31",sheet)
    fBold("C33",sheet)
    fBold("D3",sheet)
    fBold("D7",sheet)
    fBold("D10",sheet)
    fBold("D18",sheet)
    fBold("D22",sheet)
    fBold("D27",sheet)
    fBold("D31",sheet)
    fBold("D33",sheet)
    fBold("E3",sheet)
    fBold("E7",sheet)
    fBold("E10",sheet)
    fBold("E18",sheet)
    fBold("E22",sheet)
    fBold("E27",sheet)
    fBold("E31",sheet)
    fBold("E33",sheet)
    fBold("G3",sheet)
    fBold("G7",sheet)
    fBold("G10",sheet)
    fBold("G18",sheet)
    fBold("G22",sheet)
    fBold("G27",sheet)
    fBold("G31",sheet)
    fBold("G33",sheet)

#compile multimodal course enrollment from UM)SR_STRMTX_CLASSES_xxxx.xls
def compCourse(xls,sheet,sem):
    if xls == "":
        return
    xls2xlsx(xls+".xls",xls+".xlsx")
    courseWB=loadWorkBook(xls+".xlsx")
    wb=courseWB.active

    col0=wb['G'] #grad or undergrad?
    col0=fKeys(col0)
    cnt0=len(col0)
    temp01=cntU("UGRD",col0,0,cnt0)
    temp02=cntU("GRAD",col0,0,cnt0)

    col1=wb['P']
    col1=fKeys(col1)
    col1.remove("Facil ID")
    temp11= findU(col1)

    col2=wb['E']
    col2=fKeys(col2)
    col2.remove("Subject")
    temp21=findU(col2)

    sheet[sem+'3']=temp01+temp02
    sheet[sem+'4']=temp01
    sheet[sem+'5']=temp02
    sheet[sem+'9']=temp11
    sheet[sem+'10']=temp21

#compile uber transportation enrollment from Uber_xxxx_Costs&Calcs.csv
def compUber(csv,sheet,sem):
    if csv == "":
        return
    wb=C2E(csv+".csv")
    col0=wb['N'] #number of people as emails
    col0=fKeys(col0)
    col0.remove("Email")
    temp0= findU(col0)

    col1=wb['AB']
    col1=fKeys(col1)
    col2=wb['AK']
    col2=fKeys(col2)
    temp1=findSumU(col1,col2)

    sheet[sem+'17']=temp0
    sheet[sem+'19']=temp1
    sheet[sem+'18']=temp1/temp0

#compile event, rental and lease revenue from N/A
def compEvent():
    print("")

#compile campus based course enrollment from UM_SR_STRMTX_STUD_ENROLLED_xxxx.xls
def compEnroll(xls,sheet,sem):
    if xls == "":
        return
    xls2xlsx(xls+".xls",xls+".xlsx")
    enrollWB=loadWorkBook(xls+".xlsx")
    wb=enrollWB.active

    col0=wb['C']
    col0=fKeys(col0)
    col0.remove("Class Nbr")
    col1=wb['Q']
    col1=fKeys(col1)
    col1.remove("Tot Enrl")  
    temp0=findSumE(col0,col1)

    sheet[sem+'27']=temp0
    sheet[sem+'28']=temp0

#compile residential student population (unique residents)from STRPLMTX_ResList_xxxx.xls
def compResident(xls,sheet,sem):
    if xls == "":
        return
    xls2xlsx(xls+".xls",xls+".xlsx")
    resWB=loadWorkBook(xls+".xlsx")
    wb=resWB.active
    sheet[sem+'31']=int(wb['B1'].value)

#compile courses offered (copy from multi modal course enrollment total)
def compOffer(sheet,sem):
    sheet[sem+'33']=sheet[sem+'3'].value
    sheet[sem+'34']=sheet[sem+'3'].value

#compile total
def compileTotal(sheet):
    sheet["G3"] = (sheet['C3'].value or 0)+(sheet['D3'].value or 0)+(sheet['E3'].value or 0)
    sheet["G4"] = (sheet['C4'].value or 0)+(sheet['D4'].value or 0)+(sheet['E4'].value or 0)
    sheet["G5"] = (sheet['C5'].value or 0)+(sheet['D5'].value or 0)+(sheet['E5'].value or 0)
    sheet["G9"] = (sheet['C9'].value or 0)+(sheet['D9'].value or 0)+(sheet['E9'].value or 0)
    sheet["G10"] =(sheet['C10'].value or 0)+(sheet['D10'].value or 0)+(sheet['E10'].value or 0)
    sheet["G17"] =(sheet['C17'].value or 0)+(sheet['D17'].value or 0)+(sheet['E17'].value or 0)
    sheet["G18"] =(sheet['C18'].value or 0)+(sheet['D18'].value or 0)+(sheet['E18'].value or 0)
    sheet["G19"] =(sheet['C19'].value or 0)+(sheet['D19'].value or 0)+(sheet['E19'].value or 0)
    sheet["G27"] =(sheet['C27'].value or 0)+(sheet['D27'].value or 0)+(sheet['E27'].value or 0)
    sheet["G28"] =(sheet['C28'].value or 0)+(sheet['D28'].value or 0)+(sheet['E28'].value or 0)
    sheet["G31"] =(sheet['C31'].value or 0)+(sheet['D31'].value or 0)+(sheet['E31'].value or 0)
    sheet["G33"] =(sheet['C33'].value or 0)+(sheet['D33'].value or 0)+(sheet['E33'].value or 0)
    sheet["G34"] =(sheet['C34'].value or 0)+(sheet['D34'].value or 0)+(sheet['E34'].value or 0)
#export excel
def export(name):
    name.save(filename="Test.xlsx")



#MAIN
print("Test")
ROOT = tk.Tk()
ROOT.withdraw()
# the input dialog
year=simpledialog.askstring(title="Test",prompt="What Academic year is this?")
course0 = simpledialog.askstring(title="Test",prompt="Enter Summer file: UM_SR_STRMTX_CLASSES_xxxx")
uber0=simpledialog.askstring(title="Test",prompt="Enter Summer file: uber_straplamtx_xxxxxxxxx")
enroll0=simpledialog.askstring(title="Test",prompt="Enter Summer file: UM_SR_STRMTX_STUD_ENROLLED_xxxx")
resident0=simpledialog.askstring(title="Test",prompt="Enter Summer file: STRPLMTX_ResList_xxxx")
course1 = simpledialog.askstring(title="Test",prompt="Enter Fall file: UM_SR_STRMTX_CLASSES_xxxx")
uber1=simpledialog.askstring(title="Test",prompt="Enter Fall file: uber_straplamtx_xxxxxxxxx")
enroll1=simpledialog.askstring(title="Test",prompt="Enter Fall file: UM_SR_STRMTX_STUD_ENROLLED_xxxx")
resident1=simpledialog.askstring(title="Test",prompt="Enter Fall file: STRPLMTX_ResList_xxxx")
course2 = simpledialog.askstring(title="Test",prompt="Enter Spring file: UM_SR_STRMTX_CLASSES_xxxx")
uber2=simpledialog.askstring(title="Test",prompt="Enter Spring file: uber_straplamtx_xxxxxxxxx")
enroll2=simpledialog.askstring(title="Test",prompt="Enter Spring file: UM_SR_STRMTX_STUD_ENROLLED_xxxx")
resident2=simpledialog.askstring(title="Test",prompt="Enter Spring file: STRPLMTX_ResList_xxxx")


WB=newWorkBook()
sheet=makeSheet(WB)
foundation(sheet, year)
compCourse(course0,sheet, 'C')
compUber(uber0,sheet, 'C')
compEnroll(enroll0,sheet,'C')
compResident(resident0,sheet,'C')
compOffer(sheet,'C')


compCourse(course1,sheet, 'D')
compUber(uber1,sheet, 'D')
compEnroll(enroll1,sheet,'D')
compResident(resident1,sheet,'D')
compOffer(sheet,'D')


compCourse(course2,sheet, 'E')
compUber(uber2,sheet, 'E')
compEnroll(enroll2,sheet,'E')
compResident(resident2,sheet,'E')
compOffer(sheet,'E')

compileTotal(sheet)
export(WB)
messagebox.showinfo("Finish","Workbook Test.xlsx has been created")
